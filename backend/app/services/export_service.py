import json
from datetime import datetime
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class ExportService:
    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(exist_ok=True)
    
    def create_meeting_output_dir(self) -> Path:
        """
        为单次会议创建独立输出目录，避免覆盖历史导出文件。
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        short_id = uuid4().hex[:6]
        meeting_dir = self.output_dir / f"meeting_{timestamp}_{short_id}"
        meeting_dir.mkdir(parents=True, exist_ok=True)
        return meeting_dir    

    def export_json(self, meeting_result: dict[str, Any]) -> Path:
        """
        导出 JSON 文件，主要用于调试、系统集成和后续 API 返回。
        """
        output_path = self.output_dir / "meeting_result.json"

        output_path.write_text(
            json.dumps(meeting_result, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        return output_path
    
    def export_json_to_dir(self, meeting_result: dict[str, Any], target_dir: Path) -> Path:
        output_path = target_dir / "meeting_result.json"

        output_path.write_text(
            json.dumps(meeting_result, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        return output_path    

    def export_markdown(self, meeting_result: dict[str, Any]) -> Path:
        """
        导出 Markdown 会议纪要，适合复制到飞书文档、Notion、Word 等。
        """
        overview = meeting_result.get("meeting_overview", {})
        decisions = meeting_result.get("decisions", [])
        action_items = meeting_result.get("action_items", [])
        risks = meeting_result.get("risks", [])

        meeting_title = overview.get("meeting_title", "unknown")
        meeting_type = overview.get("meeting_type", "unknown")
        meeting_date = overview.get("meeting_date", "unknown")
        duration = overview.get("duration", "unknown")
        participants = overview.get("participants", [])
        main_topics = overview.get("main_topics", [])
        summary = meeting_result.get("summary", "")
        transcript_text = meeting_result.get("transcript_text", "")

        lines = []

        lines.append(f"# 会议纪要：{meeting_title}")
        lines.append("")
        lines.append("## 一、会议概览")
        lines.append("")
        lines.append(f"- 会议类型：{meeting_type}")
        lines.append(f"- 会议日期：{meeting_date}")
        lines.append(f"- 会议时长：{duration}")
        lines.append(f"- 参与人：{', '.join(participants) if participants else 'unknown'}")
        lines.append(f"- 主要议题：{', '.join(main_topics) if main_topics else 'unknown'}")
        lines.append("")

        lines.append("## 二、会议摘要")
        lines.append("")
        lines.append(summary or "unknown")
        lines.append("")

        lines.append("## 三、关键决策")
        lines.append("")
        lines.append("| 决策 | 负责人 | 证据时间 | 置信度 |")
        lines.append("|---|---|---|---|")

        if decisions:
            for item in decisions:
                lines.append(
                    f"| {item.get('decision', 'unknown')} "
                    f"| {item.get('owner', 'unknown')} "
                    f"| {item.get('evidence_time', 'unknown')} "
                    f"| {item.get('confidence', 'unknown')} |"
                )
        else:
            lines.append("| 无明确关键决策 | unknown | unknown | unknown |")

        lines.append("")

        lines.append("## 四、TODO 清单")
        lines.append("")
        lines.append("| 任务 | 责任人 | 截止时间 | 优先级 | 状态 | 证据时间 |")
        lines.append("|---|---|---|---|---|---|")

        if action_items:
            for item in action_items:
                lines.append(
                    f"| {item.get('task', 'unknown')} "
                    f"| {item.get('owner', 'unknown')} "
                    f"| {item.get('deadline', 'unknown')} "
                    f"| {item.get('priority', 'unknown')} "
                    f"| {item.get('status', 'unknown')} "
                    f"| {item.get('evidence_time', 'unknown')} |"
                )
        else:
            lines.append("| 无明确 TODO | unknown | unknown | unknown | unknown | unknown |")

        lines.append("")

        lines.append("## 五、风险与待确认问题")
        lines.append("")
        lines.append("| 风险 | 影响 | 建议跟进 | 负责人 | 严重程度 | 证据时间 |")
        lines.append("|---|---|---|---|---|---|")

        if risks:
            for item in risks:
                lines.append(
                    f"| {item.get('risk', 'unknown')} "
                    f"| {item.get('impact', 'unknown')} "
                    f"| {item.get('suggested_followup', 'unknown')} "
                    f"| {item.get('owner', 'unknown')} "
                    f"| {item.get('severity', 'unknown')} "
                    f"| {item.get('evidence_time', 'unknown')} |"
                )
        else:
            lines.append("| 无明确风险 | unknown | unknown | unknown | unknown | unknown |")

        lines.append("")
        lines.append("## 六、原始转写文本")
        lines.append("")

        if transcript_text:
            lines.append("```text")
            lines.append(transcript_text)
            lines.append("```")
        else:
            lines.append("无原始转写文本")

        markdown_text = "\n".join(lines)

        output_path = self.output_dir / "meeting_minutes.md"
        output_path.write_text(markdown_text, encoding="utf-8")

        return output_path
    
    def export_markdown_to_dir(self, meeting_result: dict[str, Any], target_dir: Path) -> Path:
        original_output_dir = self.output_dir
        self.output_dir = target_dir

        try:
            return self.export_markdown(meeting_result)
        finally:
            self.output_dir = original_output_dir    

    def export_excel(self, meeting_result: dict[str, Any]) -> Path:
        """
        导出 Excel 文件。
        一个 Excel 文件包含：
        1. Meeting Overview
        2. Decisions
        3. Action Items
        4. Risks
        """
        output_path = self.output_dir / "meeting_result.xlsx"

        wb = Workbook()

        default_sheet = wb.active
        default_sheet.title = "Meeting Overview"

        self._write_overview_sheet(default_sheet, meeting_result)
        self._write_decisions_sheet(wb, meeting_result)
        self._write_action_items_sheet(wb, meeting_result)
        self._write_risks_sheet(wb, meeting_result)
        self._write_transcript_sheet(wb, meeting_result)

        wb.save(output_path)

        return output_path
    
    def export_excel_to_dir(self, meeting_result: dict[str, Any], target_dir: Path) -> Path:
        original_output_dir = self.output_dir
        self.output_dir = target_dir

        try:
            return self.export_excel(meeting_result)
        finally:
            self.output_dir = original_output_dir

    def export_all(self, meeting_result: dict[str, Any]) -> dict[str, Path]:
        """
        一次性导出所有格式。
        """
        return {
            "json": self.export_json(meeting_result),
            "markdown": self.export_markdown(meeting_result),
            "excel": self.export_excel(meeting_result),
        }

    def _style_header_row(self, ws):
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

    def _auto_adjust_columns(self, ws):
        for column_cells in ws.columns:
            max_length = 0
            column_index = column_cells[0].column
            column_letter = get_column_letter(column_index)

            for cell in column_cells:
                value = cell.value
                if value is None:
                    continue

                value_length = len(str(value))
                if value_length > max_length:
                    max_length = value_length

            adjusted_width = min(max_length + 4, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _write_overview_sheet(self, ws, meeting_result: dict[str, Any]):
        overview = meeting_result.get("meeting_overview", {})

        rows = [
            ["字段", "内容"],
            ["会议标题", overview.get("meeting_title", "unknown")],
            ["会议类型", overview.get("meeting_type", "unknown")],
            ["会议日期", overview.get("meeting_date", "unknown")],
            ["会议时长", overview.get("duration", "unknown")],
            ["参与人", ", ".join(overview.get("participants", []))],
            ["主要议题", ", ".join(overview.get("main_topics", []))],
            ["会议摘要", meeting_result.get("summary", "unknown")],
        ]

        for row in rows:
            ws.append(row)

        self._style_header_row(ws)
        self._auto_adjust_columns(ws)

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _write_decisions_sheet(self, wb, meeting_result: dict[str, Any]):
        ws = wb.create_sheet("Decisions")

        ws.append(["决策", "负责人", "证据时间", "证据片段ID", "置信度"])

        decisions = meeting_result.get("decisions", [])

        if decisions:
            for item in decisions:
                ws.append([
                    item.get("decision", "unknown"),
                    item.get("owner", "unknown"),
                    item.get("evidence_time", "unknown"),
                    ", ".join(item.get("evidence_segment_ids", [])),
                    item.get("confidence", "unknown"),
                ])
        else:
            ws.append(["无明确关键决策", "unknown", "unknown", "", "unknown"])

        self._style_header_row(ws)
        self._auto_adjust_columns(ws)

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _write_action_items_sheet(self, wb, meeting_result: dict[str, Any]):
        ws = wb.create_sheet("Action Items")

        ws.append(["任务", "责任人", "截止时间", "优先级", "状态", "来源", "证据时间", "证据片段ID"])

        action_items = meeting_result.get("action_items", [])

        if action_items:
            for item in action_items:
                ws.append([
                    item.get("task", "unknown"),
                    item.get("owner", "unknown"),
                    item.get("deadline", "unknown"),
                    item.get("priority", "unknown"),
                    item.get("status", "unknown"),
                    item.get("source", "unknown"),
                    item.get("evidence_time", "unknown"),
                    ", ".join(item.get("evidence_segment_ids", [])),
                ])
        else:
            ws.append(["无明确 TODO", "unknown", "unknown", "unknown", "unknown", "unknown", "unknown", ""])

        self._style_header_row(ws)
        self._auto_adjust_columns(ws)

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _write_risks_sheet(self, wb, meeting_result: dict[str, Any]):
        ws = wb.create_sheet("Risks")

        ws.append(["风险", "影响", "建议跟进", "负责人", "严重程度", "证据时间", "证据片段ID"])

        risks = meeting_result.get("risks", [])

        if risks:
            for item in risks:
                ws.append([
                    item.get("risk", "unknown"),
                    item.get("impact", "unknown"),
                    item.get("suggested_followup", "unknown"),
                    item.get("owner", "unknown"),
                    item.get("severity", "unknown"),
                    item.get("evidence_time", "unknown"),
                    ", ".join(item.get("evidence_segment_ids", [])),
                ])
        else:
            ws.append(["无明确风险", "unknown", "unknown", "unknown", "unknown", "unknown", ""])

        self._style_header_row(ws)
        self._auto_adjust_columns(ws)

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    def _write_transcript_sheet(self, wb, meeting_result: dict[str, Any]):
        ws = wb.create_sheet("Transcript")

        transcript_segments = meeting_result.get("transcript_segments", [])
        transcript_text = meeting_result.get("transcript_text", "")

        if transcript_segments:
            ws.append(["片段ID", "开始时间", "结束时间", "发言人", "文本"])

            for item in transcript_segments:
                ws.append([
                    item.get("id", ""),
                    item.get("start", ""),
                    item.get("end", ""),
                    item.get("speaker", "unknown"),
                    item.get("text", ""),
                ])
        else:
            ws.append(["原始转写文本"])
            ws.append([transcript_text or "无原始转写文本"])

        self._style_header_row(ws)
        self._auto_adjust_columns(ws)

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)    

    def export_transcript_to_dir(self, meeting_result: dict[str, Any], target_dir: Path) -> Path:
        transcript_text = meeting_result.get("transcript_text", "")
        output_path = target_dir / "transcript.txt"

        output_path.write_text(
            transcript_text or "无原始转写文本",
            encoding="utf-8",
        )

        return output_path
    
    def export_meeting_record(self, meeting_result: dict[str, Any]) -> dict[str, Path]:
        """
        导出完整会议记录到独立目录。
        """
        meeting_dir = self.create_meeting_output_dir()

        json_path = self.export_json_to_dir(meeting_result, meeting_dir)
        markdown_path = self.export_markdown_to_dir(meeting_result, meeting_dir)
        excel_path = self.export_excel_to_dir(meeting_result, meeting_dir)
        transcript_path = self.export_transcript_to_dir(meeting_result, meeting_dir)

        return {
            "meeting_dir": meeting_dir,
            "json": json_path,
            "markdown": markdown_path,
            "excel": excel_path,
            "transcript": transcript_path,
        }